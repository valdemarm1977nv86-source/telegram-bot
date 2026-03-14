import os
import asyncio
from aiohttp import web
from aiogram import Bot, Dispatcher, F
from aiogram.types import Message, ReplyKeyboardMarkup, KeyboardButton, FSInputFile, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.filters import CommandStart
from aiogram.utils.media_group import MediaGroupBuilder
from openpyxl import Workbook, load_workbook
from datetime import datetime

TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID = 1584040288

bot = Bot(token=TOKEN)
dp = Dispatcher()

excel_file = "leads.xlsx"

visitors = 0
sources = {}

# кнопка заявки
order_kb = ReplyKeyboardMarkup(
    keyboard=[[KeyboardButton(text="🛒 Оставить заявку")]],
    resize_keyboard=True
)

# кнопка отправки номера
phone_kb = ReplyKeyboardMarkup(
    keyboard=[[KeyboardButton(text="📱 Отправить номер телефона", request_contact=True)]],
    resize_keyboard=True
)

# кнопка политики
policy_kb = InlineKeyboardMarkup(
    inline_keyboard=[
        [InlineKeyboardButton(text="📜 Политика обработки данных", callback_data="policy")]
    ]
)


# загрузка описания
def load_description():

    try:
        if not os.path.exists("description.txt"):
            return "Описание товара не найдено"

        with open("description.txt", "r", encoding="utf-8") as f:
            return f.read()

    except Exception as e:
        print("Description error:", e)
        return "Ошибка загрузки описания"


# загрузка фото
def load_photos():

    photos = []

    try:

        for file in sorted(os.listdir()):
            if file.lower().endswith((".jpg", ".jpeg", ".png")):
                if os.path.isfile(file):
                    photos.append(FSInputFile(file))

    except Exception as e:
        print("Photo error:", e)

    return photos


# сохранение лида
def save_to_excel(username, phone, source):

    try:

        if not os.path.exists(excel_file):
            wb = Workbook()
            ws = wb.active
            ws.append(["Дата", "Username", "Телефон", "Источник"])
            wb.save(excel_file)

        wb = load_workbook(excel_file)
        ws = wb.active

        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            username,
            phone,
            source
        ])

        wb.save(excel_file)

    except Exception as e:
        print("Excel error:", e)


@dp.message(CommandStart())
async def start(message: Message):

    global visitors
    visitors += 1

    try:

        args = message.text.split()

        source = "unknown"

        if len(args) > 1:
            source = args[1]

        if source not in sources:
            sources[source] = 0

        sources[source] += 1

        username = message.from_user.username or message.from_user.first_name

        await bot.send_message(
            ADMIN_ID,
            f"Новый посетитель @{username}\nИсточник: {source}\nВсего: {visitors}"
        )

        photos = load_photos()

        if photos:

            media = MediaGroupBuilder()

            for photo in photos:
                media.add_photo(photo)

            await message.answer_media_group(media.build())

        text = load_description()

        await message.answer(text, reply_markup=order_kb)

    except Exception as e:
        print("Start error:", e)


# кнопка заявки
@dp.message(F.text == "🛒 Оставить заявку")
async def order(message: Message):

    await message.answer(
        "Перед отправкой номера ознакомьтесь с политикой обработки данных",
        reply_markup=policy_kb
    )


# политика
@dp.callback_query(F.data == "policy")
async def policy(callback):

    text = """
Политика обработки персональных данных

Отправляя номер телефона вы соглашаетесь
на обработку персональных данных.

Данные используются только для связи
и оформления заказа.
"""

    await callback.message.answer(text)
    await callback.message.answer("Теперь отправьте номер телефона", reply_markup=phone_kb)


# получение контакта
@dp.message(F.contact)
async def contact(message: Message):

    try:

        phone = message.contact.phone_number

        if not phone:
            await message.answer("Не удалось получить номер.")
            return

        username = message.from_user.username or message.from_user.first_name

        save_to_excel(username, phone, "telegram")

        await bot.send_message(
            ADMIN_ID,
            f"Новая заявка\n@{username}\nТелефон: {phone}"
        )

        await message.answer("Спасибо! Мы свяжемся с вами.")

    except Exception as e:
        print("Contact error:", e)


# WEB SERVER для Render
async def handle(request):
    return web.Response(text="Bot running")


async def start_web():

    app = web.Application()
    app.router.add_get("/", handle)

    port = int(os.environ.get("PORT", 10000))

    runner = web.AppRunner(app)
    await runner.setup()

    site = web.TCPSite(runner, "0.0.0.0", port)
    await site.start()


async def main():

    await asyncio.gather(
        dp.start_polling(bot),
        start_web()
    )


if __name__ == "__main__":
    asyncio.run(main())
